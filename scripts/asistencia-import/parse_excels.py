#!/usr/bin/env python3
"""
Parser v2 de los Excels de asistencia 2025 y 2026.

INSIGHTS clave:
- El nombre de la HOJA es la fuente de verdad del mes (no el label del bloque,
  que a veces dice "Enero" por error de copy/paste en hojas posteriores).
- Los números de día están en la fila "NOMBRE Y APELLIDO" (col3 a col32 o col29).
  Hay que leerlos directamente en vez de asumir col3=día1.
- El año está al final del header del bloque, pero la columna exacta varía.
  Estrategia: buscar cualquier valor numérico 2024-2030 en la fila del header.
- Códigos de celda: A, R = asistió; V, P, E, F, J = otra cosa (no registramos).

Output:
  - alumnos_excel.json : lista única de alumnos (con nombre normalizado + variantes)
  - asistencias_excel.json : [{nombre_original, fecha, tipo, programa, ...}]
  - stats_excel.json : estadísticas
"""

import openpyxl
import json
import re
import unicodedata
from pathlib import Path
from datetime import date

BASE_DIR = Path('/Users/sebastien/Documents/AmasTeamWolf/docs/Excels asistencias')
OUT_DIR = Path('/Users/sebastien/Documents/AmasTeamWolf/scripts/asistencia-import')

MESES = {
    'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4,
    'mayo': 5, 'junio': 6, 'julio': 7, 'agosto': 8,
    'setiembre': 9, 'septiembre': 9,
    'octubre': 10, 'noviembre': 11, 'diciembre': 12,
}

# Letras aceptadas como "asistió físicamente a clase"
ASISTIO_CODES = {'A', 'R'}
OTROS_CODES = {'V', 'P', 'E', 'F', 'J'}
ALL_CODES = ASISTIO_CODES | OTROS_CODES


def normalize_cell(v):
    """Devuelve la celda como string upper-cased limpio, o None."""
    if v is None:
        return None
    s = str(v).strip().upper()
    if not s:
        return None
    return s


def is_block_header_row(row):
    """col2 = nombre de mes (ignoramos 'Total'), col3 = programa."""
    col2 = normalize_cell(row[1] if len(row) > 1 else None)
    col3 = normalize_cell(row[2] if len(row) > 2 else None)
    if not col2 or not col3:
        return False
    if 'TOTAL' in col2:
        return False
    # Debe ser un mes válido (el label puede ser incorrecto pero debe ser un mes)
    return col2.lower() in MESES


def is_block_end_row(row):
    col2 = normalize_cell(row[1] if len(row) > 1 else None)
    if not col2:
        return False
    return 'TOTAL' in col2


def is_nombre_empleado_row(row):
    """
    Detecta la fila header del bloque: col2 es el label de los alumnos
    ('ALUMNOS', 'NOMBRE DEL EMPLEADO', 'NOMBRE Y APELLIDO(S)') y col3+ son días.
    Validación robusta: col3 debe ser un número entre 1 y 2 (primer día del mes).
    """
    col2 = normalize_cell(row[1] if len(row) > 1 else None)
    if not col2:
        return False
    # Variantes conocidas
    is_label = (
        col2 == 'ALUMNOS'
        or ('NOMBRE' in col2 and ('APELLIDO' in col2 or 'EMPLEADO' in col2))
    )
    if not is_label:
        return False
    # Doble check: col3 debe ser un día numérico pequeño
    col3 = row[2] if len(row) > 2 else None
    if col3 is None:
        return False
    try:
        v = float(str(col3).strip())
        return 1 <= v <= 31 and v == int(v)
    except (ValueError, TypeError):
        return False


def extract_year_from_row(row, file_fallback):
    """Busca cualquier valor numérico 2020-2030 en la fila del header del bloque."""
    for cell in row:
        if cell is None:
            continue
        try:
            v = float(str(cell).strip())
            if 2020 <= v <= 2030:
                return int(v)
        except (ValueError, TypeError):
            continue
    return file_fallback


def extract_day_mapping(header_row):
    """
    Dada la fila 'NOMBRE Y APELLIDO | 1 | 2 | ...', devuelve
    {col_1based: dia_num}. Ignora la columna 2 (label) y 'Total de días'.
    """
    mapping = {}
    for col_idx_0based, cell in enumerate(header_row):
        if col_idx_0based < 2:  # col1 o col2 (label)
            continue
        if cell is None:
            continue
        try:
            # Los días pueden venir como 1, 1.0, "1", "1.0"
            v = float(str(cell).strip())
            if 1 <= v <= 31 and v == int(v):
                mapping[col_idx_0based + 1] = int(v)
        except (ValueError, TypeError):
            # Puede ser "Total de días" u otra cosa → ignorar
            continue
    return mapping


def parse_sheet(ws, sheet_name, filename):
    """Recorre una hoja de mes y devuelve lista de marcas + stats."""
    marcas = []
    skipped_invalid_dates = 0
    skipped_block_label_mismatch = 0

    sheet_mes_num = MESES.get(sheet_name.strip().lower())
    if not sheet_mes_num:
        return marcas, {}

    # Fallback year del nombre del archivo
    year_fallback = None
    if '2025' in filename:
        year_fallback = 2025
    elif '2026' in filename:
        year_fallback = 2026

    rows = list(ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True))
    current_block = None  # {programa, year, day_map}
    i = 0
    while i < len(rows):
        row = rows[i]
        row_num = i + 1

        if is_block_header_row(row):
            # Nuevo bloque: extraer programa, year, luego leer la fila
            # "NOMBRE Y APELLIDO" para el day mapping.
            col3 = str(row[2]).strip()  # programa
            year = extract_year_from_row(row, year_fallback)

            # Buscar la fila "NOMBRE Y APELLIDO" que es +1 o +2 filas después
            day_map = {}
            for offset in range(1, 5):
                if i + offset >= len(rows):
                    break
                candidate = rows[i + offset]
                if is_nombre_empleado_row(candidate):
                    day_map = extract_day_mapping(candidate)
                    break

            if not day_map:
                # No encontramos mapeo de días → skip
                i += 1
                continue

            current_block = {
                'programa': col3,
                'year': year,
                'day_map': day_map,
                'label_mes': normalize_cell(row[1]).lower(),
            }

            # Si el label del bloque no coincide con el nombre de la hoja,
            # anotar pero usar el nombre de la hoja como verdad.
            if current_block['label_mes'] != sheet_name.strip().lower():
                skipped_block_label_mismatch += 1

            i += 1
            continue

        if is_block_end_row(row):
            current_block = None
            i += 1
            continue

        if is_nombre_empleado_row(row):
            i += 1
            continue

        # Fila de alumno dentro de un bloque activo
        if current_block is None:
            i += 1
            continue

        col2 = row[1] if len(row) > 1 else None
        if col2 is None or not str(col2).strip():
            i += 1
            continue
        nombre = str(col2).strip()
        nombre_upper = nombre.upper()

        # Ignorar metadatos que puedan haber colado
        if any(kw in nombre_upper for kw in ['NOMBRE', 'CLAVES', 'ASISTENCIA DE', 'TOTAL', 'NÚMERO']):
            i += 1
            continue

        # Extraer marcas día por día usando el day_map del bloque
        for col_1based, dia in current_block['day_map'].items():
            col_0based = col_1based - 1
            if col_0based >= len(row):
                continue
            cell = row[col_0based]
            code = normalize_cell(cell)
            if code is None:
                continue
            if code not in ALL_CODES:
                continue

            year = current_block['year']
            try:
                fecha = date(year, sheet_mes_num, dia)
            except (ValueError, TypeError):
                skipped_invalid_dates += 1
                continue

            marcas.append({
                'nombre_original': nombre,
                'fecha': fecha.isoformat(),
                'year': year,
                'mes': sheet_mes_num,
                'dia': dia,
                'programa': current_block['programa'],
                'tipo': code,
                'sheet': sheet_name,
                'file': filename,
                'label_mes_bloque': current_block['label_mes'],
            })

        i += 1

    stats = {
        'skipped_invalid_dates': skipped_invalid_dates,
        'blocks_con_label_incorrecto': skipped_block_label_mismatch,
    }
    return marcas, stats


def normalize_nombre(s):
    """Normalización agresiva: upper, sin tildes, sin dobles espacios, sin puntuación."""
    if not s:
        return ''
    s = s.strip().upper()
    s = ''.join(c for c in unicodedata.normalize('NFD', s)
                if unicodedata.category(c) != 'Mn')
    s = s.replace('Ñ', 'N')
    s = re.sub(r'\s+', ' ', s)
    s = re.sub(r'[^A-Z0-9 ]', '', s)
    return s.strip()


def main():
    files = [
        'Asistencia 2025 (1).xlsx',
        'Asistencia 2026 (1).xlsx',
    ]

    all_marcas = []
    file_stats = {}

    for fname in files:
        path = BASE_DIR / fname
        print(f'\n{"="*80}\nProcessing: {fname}\n{"="*80}')
        wb = openpyxl.load_workbook(path, data_only=True)

        file_stats[fname] = {'sheets': {}, 'total_marcas': 0}

        for sheet_name in wb.sheetnames:
            if sheet_name.strip().lower() not in MESES:
                continue
            ws = wb[sheet_name]
            marcas, stats = parse_sheet(ws, sheet_name, fname)
            file_stats[fname]['sheets'][sheet_name] = {
                'marcas': len(marcas),
                **stats,
            }
            file_stats[fname]['total_marcas'] += len(marcas)
            tipos = {}
            for m in marcas:
                tipos[m['tipo']] = tipos.get(m['tipo'], 0) + 1
            print(f'  {sheet_name:12s}: {len(marcas):5d} marcas   tipos={tipos}   mismatches={stats.get("blocks_con_label_incorrecto", 0)}')
            all_marcas.extend(marcas)

    # Dedup alumnos
    alumnos_set = {}
    for a in all_marcas:
        key = normalize_nombre(a['nombre_original'])
        if not key:
            continue
        if key not in alumnos_set:
            alumnos_set[key] = {
                'normalized': key,
                'variantes': set(),
                'num_marcas': 0,
                'num_asistencias': 0,
                'programas': set(),
                'años': set(),
            }
        alumnos_set[key]['variantes'].add(a['nombre_original'])
        alumnos_set[key]['num_marcas'] += 1
        if a['tipo'] in ASISTIO_CODES:
            alumnos_set[key]['num_asistencias'] += 1
        alumnos_set[key]['programas'].add(a['programa'])
        alumnos_set[key]['años'].add(a['year'])

    alumnos_list = []
    for k, v in sorted(alumnos_set.items()):
        alumnos_list.append({
            'normalized': v['normalized'],
            'variantes': sorted(v['variantes']),
            'num_marcas': v['num_marcas'],
            'num_asistencias': v['num_asistencias'],
            'programas': sorted(v['programas']),
            'años': sorted(v['años']),
        })

    # Stats finales
    print(f'\n\n{"="*80}\nRESUMEN FINAL\n{"="*80}')
    print(f'Total de marcas (todas): {len(all_marcas)}')
    asistio_count = sum(1 for a in all_marcas if a['tipo'] in ASISTIO_CODES)
    print(f'Marcas ASISTIÓ (A o R): {asistio_count}')
    print(f'Alumnos únicos (normalizados): {len(alumnos_list)}')

    tipos_global = {}
    for a in all_marcas:
        tipos_global[a['tipo']] = tipos_global.get(a['tipo'], 0) + 1
    print(f'Distribución tipos: {tipos_global}')

    años_global = {}
    for a in all_marcas:
        años_global[a['year']] = años_global.get(a['year'], 0) + 1
    print(f'Distribución años: {años_global}')

    # Guardar
    with open(OUT_DIR / 'asistencias_excel.json', 'w', encoding='utf-8') as f:
        json.dump(all_marcas, f, ensure_ascii=False, indent=2)
    print(f'\nEscrito: asistencias_excel.json ({len(all_marcas)} marcas)')

    with open(OUT_DIR / 'alumnos_excel.json', 'w', encoding='utf-8') as f:
        json.dump(alumnos_list, f, ensure_ascii=False, indent=2)
    print(f'Escrito: alumnos_excel.json ({len(alumnos_list)} alumnos únicos)')

    with open(OUT_DIR / 'stats_excel.json', 'w', encoding='utf-8') as f:
        json.dump({
            'total_marcas': len(all_marcas),
            'total_asistio': asistio_count,
            'alumnos_unicos': len(alumnos_list),
            'tipos': tipos_global,
            'años': años_global,
            'por_archivo': file_stats,
        }, f, ensure_ascii=False, indent=2)
    print('Escrito: stats_excel.json')


if __name__ == '__main__':
    main()
